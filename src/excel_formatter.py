import openpyxl
import tkinter as tk
import tkinter.font as tkFont  # tkinter'ın font modülünü içe aktar
from openpyxl.styles import Border, Side
import inquirer
import sys
import os
from tkinter import Tk
from tkinter.filedialog import askopenfilename

def format_excel(file_path):
    # Kullanıcıdan inputları alma
    questions = [
        inquirer.Confirm(
            'adjust_size',
            message="Yazıların genişlik ve yükseklik ayarlarını yapmak ister misiniz?",
            default=True
        ),
        inquirer.Confirm(
            'add_border',
            message="Satırlara kenar eklemek ister misiniz?",
            default=True
        ),
    ]

    answers = inquirer.prompt(questions)

    # Eğer kullanıcı kenar eklemek istiyorsa, hangi türde olacağını sor
    border_style = None
    if answers['add_border']:
        border_options = {
            'thin': 'İnce çizgi',
            'medium': 'Orta kalınlıkta çizgi',
            'dashed': 'Kesik çizgili kenar',
            'dotted': 'Noktalı kenar',
            'thick': 'Kalın çizgi',
            'double': 'Çift çizgi',
            'hair': 'Çok ince çizgi (saç teli gibi)',
            'mediumDashed': 'Orta kalınlıkta kesik çizgili kenar',
            'dashDot': 'Tek kesik çizgi, nokta (— · — · şeklinde)',
            'mediumDashDot': 'Orta kalınlıkta kesik çizgi, nokta (— · — · şeklinde)',
            'dashDotDot': 'Çift kesik çizgi, nokta (— ·· — ·· şeklinde)',
            'mediumDashDotDot': 'Orta kalınlıkta çift kesik çizgi, nokta (— ·· — ·· şeklinde)',
            'slantDashDot': 'Çapraz kesik çizgi ve nokta',
        }

        # Kullanıcıdan border stilini seçmesini isteme
        border_choice = inquirer.list_input(
            "Kenar stili seçin",
            choices=[f"{key}: {value}" for key, value in border_options.items()]
        )
        
        # Seçilen kenar stilini key olarak ayarla
        border_style = border_choice.split(":")[0]

        print(f"Seçilen kenar stili {border_style} istediğiniz tasarımlar yapılıyor bekleyiniz.")

    # Excel dosyasını yükle
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active

    # Ekran genişliğini al ve birkaç piksel kıs
    padding = 90  # Sağdan ve soldan toplamda 90 piksel kıs
    root = tk.Tk()
    screen_width = root.winfo_screenwidth() - padding

    # Eğer boyut ayarlamaları yapılacaksa fontu ayarla
    font = None
    if answers['adjust_size']:
        font = tkFont.Font(family="Arial", size=10)  # Doğru font çağrısı

    # Sütun sayısı
    column_count = sheet.max_column
    column_width = screen_width / column_count

    # Sütun genişliklerini ayarla (boyut ayarlamaları istendiyse)
    if answers['adjust_size']:
        for col in range(1, column_count + 1):
            sheet.column_dimensions[openpyxl.utils.get_column_letter(col)].width = column_width / 7  # openpyxl'de genişlik birimi farklıdır

    # Güvenlik katsayısı
    safety_factor = 0.9  # Metin genişliğini biraz düşük tutmak için

    # Satır yüksekliğini ve hücre sarmalamayı (wrap text) ayarlama
    for row in sheet.iter_rows():
        max_height = 15  # Varsayılan satır yüksekliği (piksel cinsinden)
        for cell in row:
            if cell.value and answers['adjust_size']:
                text = str(cell.value)
                # Metindeki her karakterin genişliğini hesapla
                text_pixel_width = sum(font.measure(char) for char in text)

                # Hücreye sığabilecek maksimum piksel genişliğini güvenlik katsayısıyla azalt
                max_pixel_per_line = column_width * safety_factor

                # Gerekli satır sayısını hesapla
                num_lines = (text_pixel_width // max_pixel_per_line) + 1

                # Her satır için yaklaşık 15 piksel yüksekliği ayarla
                height = num_lines * 15

                # Eğer hesaplanan yükseklik daha büyükse max_height'i güncelle
                if height > max_height:
                    max_height = height

                # Hücredeki metin için kelime sarmalamayı (wrap_text) etkinleştir
                cell.alignment = openpyxl.styles.Alignment(wrap_text=True)

        # Satırın yüksekliğini max_height'e göre ayarla
        sheet.row_dimensions[row[0].row].height = max_height

        # Eğer kenar eklenmek isteniyorsa
        if answers['add_border']:
            # Kenar stillerini ayarla
            border = Border(
                top=Side(border_style=border_style),
                bottom=Side(border_style=border_style),
                left=Side(border_style=border_style),
                right=Side(border_style=border_style)
            )
            
            # Sadece satırın dış kenarlarına kalın kenarlık ekle
            first_cell = row[0]  # Satırın ilk hücresi
            last_cell = row[-1]  # Satırın son hücresi

            # İlk hücrenin sol, üst ve alt kenarlarına kenarlık ekle
            first_cell.border = Border(
                left=Side(border_style=border_style),
                top=Side(border_style=border_style),
                bottom=Side(border_style=border_style)
            )

            # Son hücrenin sağ, üst ve alt kenarlarına kenarlık ekle
            last_cell.border = Border(
                right=Side(border_style=border_style),
                top=Side(border_style=border_style),
                bottom=Side(border_style=border_style)
            )

            # Diğer hücrelere sadece üst ve alt kenarlarına kalın kenarlık ekle
            for cell in row[1:-1]:
                cell.border = Border(
                    top=Side(border_style=border_style),
                    bottom=Side(border_style=border_style)
                )

    # Pencereyi ölçüm işleminden sonra kapat
    root.destroy()

    # Değişiklikleri kaydet
    workbook.save(file_path)
    os.startfile(file_path)

def select_excel_file():
    print("Lütfen bir Excel dosyası seçiniz...")
    Tk().withdraw()  # Tk arayüzünü gizle
    file_path = askopenfilename(filetypes=[("Excel files", "*.xlsx")])  # Sadece .xlsx uzantılı dosyalar gösterilir
    return file_path

if __name__ == "__main__":
    # Komut satırı argümanlarını kontrol et
    if len(sys.argv) < 2:
        print("Kullanım: python3 excel_formatter.py <excel_file.xlsx>")
        sys.exit(1)
    
    file_path = sys.argv[1]

    # Dosyanın var olup olmadığını kontrol et
    if not os.path.isfile(file_path):
        print(f"Hata: {file_path} bulunamadı.")
        # Dosya bulunamadıysa tkinter ile dosya seç
        file_path = select_excel_file()

    # Dosyanın uzantısını kontrol et
    if not file_path.endswith('.xlsx'):
        print("Hata: Seçilen dosya bir Excel dosyası değil.")
        # Yanlış dosya seçildiyse tkinter ile dosya seç
        file_path = select_excel_file()

    # Dosya var ve uzantı doğruysa formatlama işlemini başlat
    try:
        format_excel(file_path)
        print(f"{file_path} başarıyla biçimlendirildi.")
    except Exception as e:
        print(f"Biçimlendirme sırasında bir hata oluştu: {e}")
